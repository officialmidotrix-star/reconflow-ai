"""
Unit tests for ReportService.

Analysis, Discrepancy, DiscrepancyRun, and AIInsight rows are constructed
directly, same testing-boundary rationale as every downstream module so
far. Uses a real LocalEncryptedFileStorage (temp dir) so the CSV/XLSX
round-trip through actual encryption, not a mock.
"""

from __future__ import annotations

import csv
import io
from datetime import date
from decimal import Decimal

import pytest
from cryptography.fernet import Fernet
from sqlalchemy import Column, String, Table, create_engine, select
from sqlalchemy.orm import Session

from modules.ai_insights.models import AIInsight
from modules.analysis_orchestration.models import Analysis, AnalysisStatus
from modules.discrepancies.models import (
    Discrepancy,
    DiscrepancyCategory,
    DiscrepancyRun,
    DiscrepancyStatus,
    Severity,
)
from modules.imports.models import Base
from modules.matching.models import ReconciliationMatch  # noqa: F401 - registers reconciliation_matches
from modules.normalization.models import Transaction  # noqa: F401 - registers transactions
from modules.reporting.dependencies import InMemoryAuditLogger
from modules.reporting.exceptions import AnalysisNotCompletedError, AnalysisNotFoundError, ReportNotFoundError
from modules.reporting.models import Report, ReportFormat
from modules.reporting.service import ReportService
from storage.file_storage import LocalEncryptedFileStorage

BRANCH_ID = "branch-1"
USER_ID = "user-1"

# "branches" is now a real table (Organization & Branch Management) -
# importing its model registers it, replacing the stand-in this file
# used before that module existed.
from modules.organizations.models import Branch  # noqa: E402,F401
# "users" is now a real table (Identity & Access) - importing its model
# registers it, replacing the stand-in this file used before that module
# existed.
from modules.identity_access.models import User  # noqa: E402,F401


@pytest.fixture()
def db(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path}/test.db")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


@pytest.fixture()
def storage(tmp_path):
    return LocalEncryptedFileStorage(tmp_path / "files", Fernet.generate_key())


@pytest.fixture()
def audit_logger():
    return InMemoryAuditLogger()


@pytest.fixture()
def service(db, storage, audit_logger):
    return ReportService(db=db, storage=storage, audit_logger=audit_logger)


def _make_analysis(db: Session, *, status: AnalysisStatus = AnalysisStatus.COMPLETED) -> Analysis:
    analysis = Analysis(
        branch_id=BRANCH_ID,
        created_by=USER_ID,
        version=1,
        status=status,
        period_start=date(2026, 1, 1),
        period_end=date(2026, 1, 31),
    )
    db.add(analysis)
    db.commit()
    db.refresh(analysis)
    return analysis


def _make_discrepancy(
    db: Session, *, analysis_id: str, category=DiscrepancyCategory.MISSING_SETTLEMENT,
    severity=Severity.LOW, estimated_loss=Decimal("10.00"),
) -> Discrepancy:
    d = Discrepancy(
        analysis_id=analysis_id,
        reconciliation_match_id="match-1",
        category=category,
        severity=severity,
        estimated_loss=estimated_loss,
    )
    db.add(d)
    db.commit()
    db.refresh(d)
    return d


def _make_discrepancy_run(db: Session, *, analysis_id: str, **counts) -> DiscrepancyRun:
    defaults = dict(total_count=1, critical_count=0, high_count=0, medium_count=0, low_count=1)
    defaults.update(counts)
    run = DiscrepancyRun(analysis_id=analysis_id, status=DiscrepancyStatus.COMPLETED, **defaults)
    db.add(run)
    db.commit()
    db.refresh(run)
    return run


def _make_ai_insight(db: Session, *, analysis_id: str, executive_summary: str) -> AIInsight:
    insight = AIInsight(
        analysis_id=analysis_id, executive_summary=executive_summary,
        provider_name="fake", model_name="fake-template-v1",
    )
    db.add(insight)
    db.commit()
    db.refresh(insight)
    return insight


class TestGenerateCsv:
    def test_generates_report_for_completed_analysis(self, service, db):
        analysis = _make_analysis(db)
        _make_discrepancy(db, analysis_id=analysis.id)

        report = service.generate_report(
            analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID
        )
        assert report.format == ReportFormat.CSV

    def test_csv_content_matches_discrepancies(self, service, db, storage):
        analysis = _make_analysis(db)
        _make_discrepancy(
            db, analysis_id=analysis.id, category=DiscrepancyCategory.INCORRECT_COMMISSION,
            severity=Severity.HIGH, estimated_loss=Decimal("250.00"),
        )

        report = service.generate_report(
            analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID
        )
        raw = storage.read(report.storage_path)
        rows = list(csv.reader(io.StringIO(raw.decode("utf-8"))))
        assert rows[0] == ["category", "severity", "estimated_loss", "reconciliation_match_id"]
        assert rows[1] == ["INCORRECT_COMMISSION", "HIGH", "250.00", "match-1"]


class TestGenerateXlsx:
    def test_includes_summary_and_detail_sheets(self, service, db, storage):
        import openpyxl

        analysis = _make_analysis(db)
        _make_discrepancy(db, analysis_id=analysis.id)
        _make_discrepancy_run(db, analysis_id=analysis.id)
        _make_ai_insight(db, analysis_id=analysis.id, executive_summary="Everything looks fine.")

        report = service.generate_report(
            analysis_id=analysis.id, format=ReportFormat.XLSX, requested_by=USER_ID
        )
        raw = storage.read(report.storage_path)
        workbook = openpyxl.load_workbook(io.BytesIO(raw))
        assert workbook.sheetnames == ["Summary", "Discrepancies"]

        summary_text = "\n".join(
            str(cell) for row in workbook["Summary"].iter_rows(values_only=True) for cell in row
        )
        assert "Everything looks fine." in summary_text

    def test_notes_missing_ai_summary(self, service, db, storage):
        import openpyxl

        analysis = _make_analysis(db)
        _make_discrepancy(db, analysis_id=analysis.id)
        # No AIInsight created for this analysis.

        report = service.generate_report(
            analysis_id=analysis.id, format=ReportFormat.XLSX, requested_by=USER_ID
        )
        raw = storage.read(report.storage_path)
        workbook = openpyxl.load_workbook(io.BytesIO(raw))
        summary_text = "\n".join(
            str(cell) for row in workbook["Summary"].iter_rows(values_only=True) for cell in row
        )
        assert "No AI summary available." in summary_text


class TestPrecondition:
    def test_rejects_non_completed_analysis(self, service, db):
        analysis = _make_analysis(db, status=AnalysisStatus.PROCESSING)
        with pytest.raises(AnalysisNotCompletedError):
            service.generate_report(
                analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID
            )

    def test_analysis_not_found_raises(self, service):
        with pytest.raises(AnalysisNotFoundError):
            service.generate_report(
                analysis_id="does-not-exist", format=ReportFormat.CSV, requested_by=USER_ID
            )


class TestDownload:
    def test_returns_original_content(self, service, db):
        analysis = _make_analysis(db)
        _make_discrepancy(db, analysis_id=analysis.id)
        report = service.generate_report(
            analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID
        )
        content, format_ = service.download_report(report_id=report.id)
        assert format_ == ReportFormat.CSV
        assert b"MISSING_SETTLEMENT" in content

    def test_nonexistent_report_raises(self, service):
        with pytest.raises(ReportNotFoundError):
            service.download_report(report_id="does-not-exist")


class TestReportsAreKeptNotSuperseded:
    def test_multiple_generations_all_persist(self, service, db):
        analysis = _make_analysis(db)
        _make_discrepancy(db, analysis_id=analysis.id)

        service.generate_report(analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID)
        service.generate_report(analysis_id=analysis.id, format=ReportFormat.XLSX, requested_by=USER_ID)
        service.generate_report(analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID)

        reports = db.execute(
            select(Report).where(Report.analysis_id == analysis.id)
        ).scalars().all()
        assert len(reports) == 3  # none superseded, unlike every other module's run output


class TestAuditLogging:
    def test_logs_generation(self, service, db, audit_logger):
        analysis = _make_analysis(db)
        _make_discrepancy(db, analysis_id=analysis.id)
        service.generate_report(analysis_id=analysis.id, format=ReportFormat.CSV, requested_by=USER_ID)
        assert audit_logger.records[-1].event == "report_generated"
        assert audit_logger.records[-1].analysis_id == analysis.id
        assert audit_logger.records[-1].metadata["format"] == "CSV"
