"""
Tabular file parsers for the Data Validation module.

Both parsers are streaming: CSV via Python's built-in `csv` module over a
decoded text stream, XLSX via openpyxl in `read_only` mode. Neither ever
materializes the whole file in memory as a table, which matters most for
XLSX - it's a zip archive containing XML, so a hostile or corrupted file
could otherwise expand far beyond its on-disk size. `max_rows` is enforced
*during* iteration so a bad file is abandoned early rather than after
being fully expanded.
"""

from __future__ import annotations

import csv
import io
from abc import ABC, abstractmethod
from collections.abc import Iterator


class FileParseError(Exception):
    """Raised when a file cannot be read as a table at all (corrupted,
    wrong internal format despite a matching extension, undecodable
    bytes, etc.) - always mapped to a FAILED validation result with an
    UNREADABLE_FILE issue, never surfaced as a raw exception to the user."""


class RowCapExceededError(Exception):
    """Raised when a file has more rows than `max_rows` - a deliberate,
    distinct signal from FileParseError so the caller can report a
    specific, actionable message rather than a generic parse failure."""

    def __init__(self, max_rows: int) -> None:
        super().__init__(f"row cap of {max_rows} exceeded")
        self.max_rows = max_rows


class TabularParser(ABC):
    @abstractmethod
    def iter_rows(self, content: bytes, *, max_rows: int) -> Iterator[list[str]]:
        """Yield each row as a list of string cell values. The first row
        yielded is the header row. Raises FileParseError if the content
        can't be read as this format at all, or RowCapExceededError if
        more than `max_rows` rows are encountered."""
        ...


class CsvParser(TabularParser):
    def iter_rows(self, content: bytes, *, max_rows: int) -> Iterator[list[str]]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise FileParseError("could not decode file as UTF-8 text") from exc

        reader = csv.reader(io.StringIO(text))
        for row_index, row in enumerate(reader):
            if row_index >= max_rows:
                raise RowCapExceededError(max_rows)
            yield row


class XlsxParser(TabularParser):
    def iter_rows(self, content: bytes, *, max_rows: int) -> Iterator[list[str]]:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - dependency always installed in practice
            raise FileParseError("xlsx support is not available") from exc

        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as exc:  # noqa: BLE001 - openpyxl raises several distinct types
            raise FileParseError("could not read file as an Excel workbook") from exc

        try:
            worksheet = workbook.worksheets[0]
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if row_index >= max_rows:
                    raise RowCapExceededError(max_rows)
                yield ["" if cell is None else str(cell) for cell in row]
        finally:
            workbook.close()


_PARSERS: dict[str, TabularParser] = {
    "csv": CsvParser(),
    "xlsx": XlsxParser(),
    "xls": XlsxParser(),
}


def parser_for_extension(extension: str) -> TabularParser:
    try:
        return _PARSERS[extension.lower()]
    except KeyError as exc:
        raise FileParseError(f"unsupported file extension: {extension}") from exc
